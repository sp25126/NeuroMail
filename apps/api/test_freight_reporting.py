import unittest
import os
import sys
import datetime
import csv
import uuid
from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool
import openpyxl

# Ensure API path is in import search path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from main import app
from database import Base, get_db
from models import (
    Tenant, User, Mailbox, FreightShipment, FreightCarrierSnapshot,
    FreightEvent, FreightAlert, FreightRawEmail, FreightTenantConfig,
    FreightReportRun, FreightReportSchedule
)
from services.report_service import generate_report, REPORTS_DIR, ensure_reports_dir
from workers.worker import process_job

# Use an isolated SQLite memory database for testing
SQLALCHEMY_DATABASE_URL = "sqlite:///:memory:"
engine = create_engine(
    SQLALCHEMY_DATABASE_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool
)
TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

def override_get_db():
    db = TestingSessionLocal()
    try:
        yield db
    finally:
        db.close()

class TestFreightReporting(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        Base.metadata.create_all(bind=engine)
        
        db = TestingSessionLocal()
        # Seed test tenants
        cls.tenant1 = Tenant(id="tenant-rep-1", name="Reporting Tenant One")
        cls.tenant2 = Tenant(id="tenant-rep-2", name="Reporting Tenant Two")
        
        # Seed test users
        cls.user1 = User(id="user-r1", email="user1@rep1.com", name="User R1", tenant_id="tenant-rep-1", role="operator")
        cls.user2 = User(id="user-r2", email="user2@rep2.com", name="User R2", tenant_id="tenant-rep-2", role="operator")
        
        # Seed test mailboxes
        cls.mailbox1 = Mailbox(id="mb-rep-1", tenant_id="tenant-rep-1", provider_type="GMAIL", connection_status="CONNECTED")
        cls.mailbox2 = Mailbox(id="mb-rep-2", tenant_id="tenant-rep-2", provider_type="OUTLOOK", connection_status="CONNECTED")
        
        # Seed Tenant-specific Freight configs
        cls.config1 = FreightTenantConfig(
            id="cfg-rep-1",
            tenant_id="tenant-rep-1",
            sync_interval_minutes=30,
            no_update_threshold_hours=24,
            storage_risk_days=3,
            freight_subject_patterns=[],
            freight_from_addresses=[],
            active_carriers=[],
            notification_email_addresses=[],
            alert_severity_threshold="medium"
        )
        cls.config2 = FreightTenantConfig(
            id="cfg-rep-2",
            tenant_id="tenant-rep-2",
            sync_interval_minutes=30,
            no_update_threshold_hours=24,
            storage_risk_days=3,
            freight_subject_patterns=[],
            freight_from_addresses=[],
            active_carriers=[],
            notification_email_addresses=[],
            alert_severity_threshold="medium"
        )
        
        db.add(cls.tenant1)
        db.add(cls.tenant2)
        db.add(cls.user1)
        db.add(cls.user2)
        db.add(cls.mailbox1)
        db.add(cls.mailbox2)
        db.add(cls.config1)
        db.add(cls.config2)
        db.commit()
        db.close()

    @classmethod
    def tearDownClass(cls):
        Base.metadata.drop_all(bind=engine)

    def setUp(self):
        app.dependency_overrides[get_db] = override_get_db
        self.client = TestClient(app)
        self.headers_t1 = {"X-Tenant-ID": "tenant-rep-1", "X-User-ID": "user-r1"}
        self.headers_t2 = {"X-Tenant-ID": "tenant-rep-2", "X-User-ID": "user-r2"}
        
        # Clear dynamic tables before each test
        db = TestingSessionLocal()
        db.query(FreightShipment).delete()
        db.query(FreightCarrierSnapshot).delete()
        db.query(FreightEvent).delete()
        db.query(FreightAlert).delete()
        db.query(FreightRawEmail).delete()
        db.query(FreightReportRun).delete()
        db.query(FreightReportSchedule).delete()
        db.commit()
        db.close()

    def test_report_generation_shipment_status_csv_and_xlsx(self):
        db = TestingSessionLocal()
        
        # Seed test shipments
        shipment = FreightShipment(
            id="sh-rep-01",
            tenant_id="tenant-rep-1",
            primary_reference="REF-XYZ",
            carrier="Maersk",
            origin_port="Shanghai",
            destination_port="Los Angeles",
            last_known_status="IN_TRANSIT",
            eta=datetime.datetime.utcnow() + datetime.timedelta(days=5),
            last_status_at=datetime.datetime.utcnow() - datetime.timedelta(hours=2),
            created_at=datetime.datetime.utcnow(),
            is_closed=False
        )
        db.add(shipment)
        db.commit()
        
        # 1. Generate CSV
        raw_data, output_uri, row_count = generate_report(
            db, "tenant-rep-1", "shipment_status", "csv"
        )
        self.assertEqual(row_count, 1)
        self.assertIsNotNone(output_uri)
        self.assertTrue("download" in output_uri)
        
        # Verify physical CSV output
        run_id = output_uri.split("/")[-1]
        file_path = os.path.join(REPORTS_DIR, f"report_shipment_status_{run_id}.csv")
        self.assertTrue(os.path.exists(file_path))
        
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            headers = next(reader)
            row = next(reader)
            self.assertEqual(headers[0], "shipment_id")
            self.assertEqual(headers[1], "primary_reference")
            self.assertEqual(row[0], "sh-rep-01")
            self.assertEqual(row[1], "REF-XYZ")
            
        # 2. Generate XLSX
        raw_data_xlsx, output_uri_xlsx, row_count_xlsx = generate_report(
            db, "tenant-rep-1", "shipment_status", "xlsx"
        )
        self.assertEqual(row_count_xlsx, 1)
        
        xlsx_run_id = output_uri_xlsx.split("/")[-1]
        file_path_xlsx = os.path.join(REPORTS_DIR, f"report_shipment_status_{xlsx_run_id}.xlsx")
        self.assertTrue(os.path.exists(file_path_xlsx))
        
        # Read back sheet using openpyxl
        wb = openpyxl.load_workbook(file_path_xlsx)
        sheet = wb.active
        self.assertEqual(sheet.cell(row=1, column=1).value, "shipment_id")
        self.assertEqual(sheet.cell(row=2, column=1).value, "sh-rep-01")
        self.assertEqual(sheet.cell(row=2, column=2).value, "REF-XYZ")
        
        # Check FreightReportRun audits
        run = db.query(FreightReportRun).filter(FreightReportRun.id == run_id).first()
        self.assertIsNotNone(run)
        self.assertEqual(run.status, "success")
        self.assertEqual(run.row_count, 1)
        self.assertEqual(run.tenant_id, "tenant-rep-1")
        
        db.close()

    def test_report_generation_aging_no_update(self):
        db = TestingSessionLocal()
        
        # Seed active shipment with older update
        now = datetime.datetime.utcnow()
        shipment = FreightShipment(
            id="sh-rep-02",
            tenant_id="tenant-rep-1",
            primary_reference="REF-AGING",
            carrier="MSC",
            origin_port="Ningbo",
            destination_port="Seattle",
            last_known_status="IN_TRANSIT",
            eta=now + datetime.timedelta(days=10),
            last_status_at=now - datetime.timedelta(hours=36), # 36 hours > 24 threshold
            created_at=now - datetime.timedelta(days=2),
            is_closed=False
        )
        db.add(shipment)
        db.commit()
        
        raw_data, output_uri, row_count = generate_report(
            db, "tenant-rep-1", "aging_no_update", "csv"
        )
        self.assertEqual(row_count, 1)
        self.assertEqual(raw_data[0]["shipment_id"], "sh-rep-02")
        self.assertEqual(raw_data[0]["no_update_breached"], True)
        db.close()

    def test_report_generation_arrival_pickup_and_quarantine(self):
        db = TestingSessionLocal()
        now = datetime.datetime.utcnow()
        
        # 1. Arrival/Pickup Shipment
        shipment = FreightShipment(
            id="sh-rep-03",
            tenant_id="tenant-rep-1",
            primary_reference="REF-ARRIVED",
            carrier="CMA CGM",
            origin_port="Yokohama",
            destination_port="Oakland",
            last_known_status="ARRIVED",
            eta=now - datetime.timedelta(days=1),
            last_status_at=now - datetime.timedelta(hours=4),
            created_at=now - datetime.timedelta(days=5),
            is_closed=False
        )
        db.add(shipment)
        
        # Snapshot marking is_arrived = True
        snap = FreightCarrierSnapshot(
            id="snap-rep-01",
            tenant_id="tenant-rep-1",
            shipment_id="sh-rep-03",
            carrier_adapter="Project44",
            reference_used="REF-ARRIVED",
            carrier_status="ARRIVED",
            is_arrived=True,
            synced_at=now - datetime.timedelta(days=4)
        )
        db.add(snap)
        
        # 2. Quarantined email
        q_email = FreightRawEmail(
            id="raw-q-01",
            tenant_id="tenant-rep-1",
            mailbox_id="mb-rep-1",
            provider="GMAIL",
            provider_message_id="msg-raw-q-01",
            subject="Damaged shipment alert",
            from_address="unknown@carrier.com",
            received_at=now - datetime.timedelta(hours=1),
            parsing_status="quarantined",
            parsing_error="No reference match found."
        )
        db.add(q_email)
        db.commit()
        
        # Run Arrival/Pickup report
        raw_arr, _, _ = generate_report(db, "tenant-rep-1", "arrival_pickup", "csv")
        self.assertEqual(len(raw_arr), 1)
        self.assertEqual(raw_arr[0]["shipment_id"], "sh-rep-03")
        self.assertEqual(raw_arr[0]["storage_risk_flag"], True)  # 4 days dwell > 3 threshold
        
        # Run Quarantine report
        raw_q, _, _ = generate_report(db, "tenant-rep-1", "quarantine", "csv")
        self.assertEqual(len(raw_q), 1)
        self.assertEqual(raw_q[0]["raw_email_id"], "raw-q-01")
        self.assertEqual(raw_q[0]["quarantine_reason"], "No reference match found.")
        
        db.close()

    def test_report_generation_failure_handling(self):
        db = TestingSessionLocal()
        with self.assertRaises(ValueError):
            generate_report(db, "tenant-rep-1", "non_existent_report_type", "csv")
            
        # Verify run recorded as failed
        run = db.query(FreightReportRun).filter(FreightReportRun.report_type == "non_existent_report_type").first()
        self.assertIsNotNone(run)
        self.assertEqual(run.status, "failed")
        self.assertIsNotNone(run.error)
        db.close()

    def test_dashboard_summary_and_historical_trend_deltas(self):
        db = TestingSessionLocal()
        now = datetime.datetime.utcnow()
        yesterday = now - datetime.timedelta(days=1)
        
        # Shipments
        s1 = FreightShipment(
            id="sh-db-1", tenant_id="tenant-rep-1", primary_reference="REF-1",
            carrier="Carrier A", origin_port="Port X", destination_port="Port Y",
            last_known_status="IN_TRANSIT", created_at=now, is_closed=False
        )
        s2 = FreightShipment(
            id="sh-db-2", tenant_id="tenant-rep-1", primary_reference="REF-2",
            carrier="Carrier B", origin_port="Port X", destination_port="Port Y",
            last_known_status="ARRIVED", created_at=yesterday - datetime.timedelta(hours=1), is_closed=False
        )
        db.add(s1)
        db.add(s2)
        
        # Snapshots
        snap1 = FreightCarrierSnapshot(
            id="snap-db-1", tenant_id="tenant-rep-1", shipment_id="sh-db-2",
            carrier_adapter="Project44", reference_used="REF-2",
            carrier_status="ARRIVED", is_arrived=True, synced_at=yesterday
        )
        db.add(snap1)
        
        # Alerts
        a1 = FreightAlert(
            id="alert-db-1", tenant_id="tenant-rep-1", shipment_id="sh-db-1",
            rule_type="ETA_BREACH", severity="critical", status="open", created_at=now,
            title="ETA Breach Alert", description="ETA breached", dedup_key="key-db-1"
        )
        a2 = FreightAlert(
            id="alert-db-2", tenant_id="tenant-rep-1", shipment_id="sh-db-1",
            rule_type="NO_UPDATE", severity="low", status="open", created_at=yesterday,
            title="No Update Alert", description="No update received", dedup_key="key-db-2"
        )
        db.add(a1)
        db.add(a2)
        
        # Quarantined
        q = FreightRawEmail(
            id="q-db-1", tenant_id="tenant-rep-1", subject="Quarantine Email",
            mailbox_id="mb-rep-1", provider="GMAIL", provider_message_id="msg-q-db-1",
            from_address="ops@company.com", parsing_status="quarantined", received_at=now
        )
        db.add(q)
        
        db.commit()
        db.close()
        
        response = self.client.get("/freight/dashboard/summary", headers=self.headers_t1)
        self.assertEqual(response.status_code, 200)
        data = response.json()
        
        self.assertEqual(data["total_shipments"], 2)
        self.assertEqual(data["shipments_arrived"], 1)
        self.assertEqual(data["quarantine_count"], 1)
        self.assertEqual(data["alerts_open_by_severity"]["critical"], 1)
        self.assertEqual(data["alerts_open_by_severity"]["low"], 1)
        self.assertEqual(data["total_shipments_delta"], 1) # s1 created now
        self.assertEqual(data["alerts_open_delta"], 1) # a1 created now, a2 was yesterday

    def test_dashboard_shipments_filtering(self):
        db = TestingSessionLocal()
        s1 = FreightShipment(
            id="sh-fl-1", tenant_id="tenant-rep-1", primary_reference="REF-FL1",
            carrier="Evergreen", origin_port="Kaohsiung", destination_port="Los Angeles",
            last_known_status="IN_TRANSIT", is_closed=False
        )
        s2 = FreightShipment(
            id="sh-fl-2", tenant_id="tenant-rep-1", primary_reference="REF-FL2",
            carrier="One Line", origin_port="Shanghai", destination_port="Oakland",
            last_known_status="ARRIVED", is_closed=False
        )
        db.add(s1)
        db.add(s2)
        db.commit()
        db.close()
        
        # Port filter
        res1 = self.client.get("/freight/dashboard/shipments?port=Shanghai", headers=self.headers_t1)
        self.assertEqual(res1.status_code, 200)
        self.assertEqual(len(res1.json()), 1)
        self.assertEqual(res1.json()[0]["id"], "sh-fl-2")
        
        # Carrier filter
        res2 = self.client.get("/freight/dashboard/shipments?carrier=Evergreen", headers=self.headers_t1)
        self.assertEqual(res2.status_code, 200)
        self.assertEqual(len(res2.json()), 1)
        
        # Status filter
        res3 = self.client.get("/freight/dashboard/shipments?status=ARRIVED", headers=self.headers_t1)
        self.assertEqual(res3.status_code, 200)
        self.assertEqual(len(res3.json()), 1)

    def test_dashboard_alerts_and_quarantine_endpoints(self):
        db = TestingSessionLocal()
        a = FreightAlert(
            id="alert-fl-1", tenant_id="tenant-rep-1", shipment_id="sh-fl-1",
            rule_type="STORAGE_RISK", severity="high", status="open", created_at=datetime.datetime.utcnow(),
            title="Storage Risk Alert", description="High storage risk", dedup_key="key-fl-1"
        )
        q = FreightRawEmail(
            id="q-fl-1", tenant_id="tenant-rep-1", subject="Oops Bad Email",
            mailbox_id="mb-rep-1", provider="GMAIL", provider_message_id="msg-q-fl-1",
            from_address="bad@carrier.com", parsing_status="quarantined", received_at=datetime.datetime.utcnow()
        )
        db.add(a)
        db.add(q)
        db.commit()
        db.close()
        
        # Alerts view endpoint
        res_alerts = self.client.get("/freight/dashboard/alerts?severity=high", headers=self.headers_t1)
        self.assertEqual(res_alerts.status_code, 200)
        self.assertEqual(len(res_alerts.json()), 1)
        self.assertEqual(res_alerts.json()[0]["id"], "alert-fl-1")
        
        # Quarantine view endpoint
        res_q = self.client.get("/freight/dashboard/quarantine", headers=self.headers_t1)
        self.assertEqual(res_q.status_code, 200)
        self.assertEqual(len(res_q.json()), 1)
        self.assertEqual(res_q.json()[0]["id"], "q-fl-1")

    def test_dashboard_shipment_detail_completeness(self):
        db = TestingSessionLocal()
        s = FreightShipment(
            id="sh-dt-1", tenant_id="tenant-rep-1", primary_reference="REF-DET",
            carrier="Maersk", origin_port="Shanghai", destination_port="Los Angeles",
            last_known_status="IN_TRANSIT"
        )
        db.add(s)
        
        snap = FreightCarrierSnapshot(
            id="snap-dt-1", tenant_id="tenant-rep-1", shipment_id="sh-dt-1",
            carrier_adapter="Project44", reference_used="REF-DET",
            carrier_status="IN_TRANSIT", synced_at=datetime.datetime.utcnow()
        )
        db.add(snap)
        
        evt = FreightEvent(
            id="evt-dt-1", tenant_id="tenant-rep-1", shipment_id="sh-dt-1",
            event_type="milestone_update", payload={"description": "Passed checkpoint"}, created_at=datetime.datetime.utcnow()
        )
        db.add(evt)
        db.commit()
        db.close()
        
        res = self.client.get("/freight/dashboard/shipments/sh-dt-1", headers=self.headers_t1)
        self.assertEqual(res.status_code, 200)
        data = res.json()
        self.assertEqual(data["shipment"]["id"], "sh-dt-1")
        self.assertEqual(len(data["snapshots"]), 1)
        self.assertEqual(len(data["events"]), 1)

    def test_secure_report_download_tenant_isolation(self):
        db = TestingSessionLocal()
        # Seed report run belonging to tenant 1
        run = FreightReportRun(
            id="run-sec-1",
            tenant_id="tenant-rep-1",
            report_type="shipment_status",
            started_at=datetime.datetime.utcnow(),
            completed_at=datetime.datetime.utcnow(),
            status="success",
            row_count=0,
            output_uri="/freight/reports/download/run-sec-1"
        )
        db.add(run)
        db.commit()
        db.close()
        
        # Ensure reports dir has the dummy physical file
        ensure_reports_dir()
        dummy_file = os.path.join(REPORTS_DIR, "report_shipment_status_run-sec-1.csv")
        with open(dummy_file, "w", newline="") as f:
            f.write("dummy,headers\n")
            
        # Download from authorized Tenant 1 -> Should work
        res_ok = self.client.get("/freight/reports/download/run-sec-1", headers=self.headers_t1)
        self.assertEqual(res_ok.status_code, 200)
        self.assertEqual(res_ok.text, "dummy,headers\n")
        
        # Download from unauthorized Tenant 2 -> Should yield 404/denied
        res_fail = self.client.get("/freight/reports/download/run-sec-1", headers=self.headers_t2)
        self.assertEqual(res_fail.status_code, 404)
        
        # Cleanup dummy file
        if os.path.exists(dummy_file):
            os.remove(dummy_file)

    def test_schedules_crud(self):
        # 1. Create schedule
        payload = {
            "report_type": "kpi_summary",
            "cron_expression": "0 8 * * *",
            "format": "xlsx",
            "recipients": ["exec@company.com"]
        }
        res = self.client.post("/freight/reports/schedules", json=payload, headers=self.headers_t1)
        self.assertEqual(res.status_code, 201)
        data = res.json()
        self.assertEqual(data["report_type"], "kpi_summary")
        self.assertEqual(data["format"], "xlsx")
        
        schedule_id = data["id"]
        
        # 2. Get schedules
        res_get = self.client.get("/freight/reports/schedules", headers=self.headers_t1)
        self.assertEqual(res_get.status_code, 200)
        self.assertEqual(len(res_get.json()), 1)
        
        # 3. Update schedule
        res_upd = self.client.put(
            f"/freight/reports/schedules/{schedule_id}",
            json={"enabled": False, "cron_expression": "30 8 * * *"},
            headers=self.headers_t1
        )
        self.assertEqual(res_upd.status_code, 200)
        self.assertEqual(res_upd.json()["enabled"], False)
        self.assertEqual(res_upd.json()["cron_expression"], "30 8 * * *")
        
        # 4. Delete schedule
        res_del = self.client.delete(f"/freight/reports/schedules/{schedule_id}", headers=self.headers_t1)
        self.assertEqual(res_del.status_code, 204)
        
        # Verify gone
        res_get_post = self.client.get("/freight/reports/schedules", headers=self.headers_t1)
        self.assertEqual(len(res_get_post.json()), 0)

    @patch("workers.worker.get_redis_client")
    @patch("database.SessionLocal")
    def test_worker_scheduled_report_locks_and_retries(self, mock_session_local, mock_redis_func):
        # Set up mocks
        mock_redis = MagicMock()
        mock_redis_func.return_value = mock_redis
        
        # Mock acquired lock
        mock_redis.set.return_value = True
        
        db_mock = MagicMock()
        mock_session_local.return_value = db_mock
        
        # Mock FreightReportSchedule record
        sched_record = FreightReportSchedule(
            id="sched-w-1",
            tenant_id="tenant-rep-1",
            report_type="kpi_summary",
            cron_expression="0 8 * * *",
            enabled=True,
            format="csv",
            created_at=datetime.datetime.utcnow()
        )
        db_mock.query().filter().filter().first.return_value = sched_record
        
        # 1. Concurrency lock collision test
        mock_redis.set.return_value = False # lock already held
        res_skip = process_job({
            "task": "freight_process_scheduled_reports",
            "args": ["tenant-rep-1", "sched-w-1"]
        })
        self.assertEqual(res_skip, "skipped")
        
        # 2. Success path test
        mock_redis.set.return_value = True
        with patch("services.report_service.generate_report") as mock_gen_report:
            res_ok = process_job({
                "task": "freight_process_scheduled_reports",
                "args": ["tenant-rep-1", "sched-w-1"]
            })
            self.assertEqual(res_ok, "success")
            mock_gen_report.assert_called_once()
            db_mock.commit.assert_called_once()
            
        # 3. Retry on failure once (first fails, second succeeds)
        db_mock.commit.reset_mock()
        mock_redis.set.return_value = True
        
        with patch("services.report_service.generate_report") as mock_gen_report_fail:
            mock_gen_report_fail.side_effect = [ValueError("First run failed"), None]
            res_retry_ok = process_job({
                "task": "freight_process_scheduled_reports",
                "args": ["tenant-rep-1", "sched-w-1"]
            })
            self.assertEqual(res_retry_ok, "success")
            self.assertEqual(mock_gen_report_fail.call_count, 2)
            db_mock.commit.assert_called_once()
            
        # 4. Final failure if both runs fail
        mock_redis.set.return_value = True
        with patch("services.report_service.generate_report") as mock_gen_report_double_fail:
            mock_gen_report_double_fail.side_effect = [ValueError("First failed"), ValueError("Retry failed")]
            res_double_fail = process_job({
                "task": "freight_process_scheduled_reports",
                "args": ["tenant-rep-1", "sched-w-1"]
            })
            self.assertEqual(res_double_fail, "failed")
            self.assertEqual(mock_gen_report_double_fail.call_count, 2)

    def test_copilot_chat_read_only_and_sources(self):
        db = TestingSessionLocal()
        s = FreightShipment(
            id="sh-cop-1", tenant_id="tenant-rep-1", primary_reference="REF-COP1",
            carrier="MSC", origin_port="Shanghai", destination_port="Los Angeles",
            last_known_status="ARRIVED", is_closed=False
        )
        db.add(s)
        # Snapshot indicating it's arrived and NOT gate_out (so at risk)
        snap = FreightCarrierSnapshot(
            id="snap-cop-1", tenant_id="tenant-rep-1", shipment_id="sh-cop-1",
            carrier_adapter="Project44", reference_used="REF-COP1",
            carrier_status="ARRIVED", is_arrived=True, synced_at=datetime.datetime.utcnow()
        )
        db.add(snap)
        db.commit()
        db.close()
        
        # Test AI assistant endpoint query
        payload = {"message": "Which shipments are at risk of storage fees?"}
        res = self.client.post("/freight/copilot/chat", json=payload, headers=self.headers_t1)
        self.assertEqual(res.status_code, 200)
        
        data = res.json()
        self.assertIsNotNone(data["response"])
        self.assertTrue(len(data["sources"]) > 0)
        # Ensure it lists the shipment as source
        self.assertEqual(data["sources"][0]["id"], "sh-cop-1")
        self.assertEqual(data["sources"][0]["ref"], "REF-COP1")
        
        # Confirm read-only constraints: no new shipments/runs are generated by asking copilot questions
        db_check = TestingSessionLocal()
        shipments_count = db_check.query(FreightShipment).count()
        self.assertEqual(shipments_count, 1) # still only 1 shipment
        db_check.close()
